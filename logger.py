import threading
import queue
import struct
from openpyxl import Workbook
from global_config import settings
from storage import SystemStorage
from file_manager import FileManager

class Logger:
    def __init__(self):
        try:
            self.file_manager = None
            self._running = False
            self.file_type = None
            self.schema = None
            self._enabled = True

            self.q = queue.Queue(maxsize=settings.QUEUE_SIZE)
            self.headers_blob = None
            self.dropped_count = 0

            self._worker = None
            self._compressor = None

            # self._compress_event = threading.Event()

        except Exception as e:
            print(f"Exception in init: {e}")

    # =========================== INITILIZER ========================
    def initialize(self, file_type: str, compress: bool):
        try:
            if not SystemStorage().checking():
                self._enabled = False
                self._running = False
                self.file_manager = None
                return   

            self.file_type = file_type or settings.DEFAULT_FILE_TYPE
            do_compress = compress if compress is not None else settings.DEFAULT_COMPRESS

            self.file_manager = FileManager(file_type=self.file_type, compress=do_compress)

        except Exception as e:
            print(f"Exception in initilizer: {e}")
            self._enabled = False

    # =========================== START ========================
    def start(self):
        try:
            if not self._enabled or self.file_manager is None:
                return   

            self._running = True

            match self.file_type:
                case "csv":
                    self._worker = threading.Thread(target=self.csv_worker, daemon=True)
                case "bin":
                    self._worker = threading.Thread(target=self.bin_worker, daemon=True)
                case "tlv.bin":
                    self._worker = threading.Thread(target=self.tlv_worker, daemon=True)
                case "xlsx":
                    self._worker = threading.Thread(target=self.xlsx_worker, daemon=True)
                case _:
                    return

            self._worker.start()

        except Exception as e:
            print(f"Exception in start: {e}")
            self._enabled = False

    # =========================== HEADER WRITER ========================
    def headers(self, *headers):
        try:
            if not self._enabled:
                return

            if not headers:
                return

            self.schema = tuple(headers)

            match self.file_type:
                case "csv":
                    self.headers_blob = (",".join(self.schema) + "\n").encode("utf-8")

                case "bin":
                    buf = bytearray()
                    buf += b"LOG1"
                    buf += (1).to_bytes(1, "little")
                    buf += len(self.schema).to_bytes(1, "little")
                    for name in self.schema:
                        b = name.encode("utf-8")
                        buf += len(b).to_bytes(1, "little")
                        buf += b
                    self.headers_blob = bytes(buf)

                case "tlv.bin":
                    buf = bytearray()
                    buf += b"TLV1"
                    buf += (1).to_bytes(1, "little")
                    buf += len(self.schema).to_bytes(1, "little")
                    FIELD_DEF = 0x01
                    for name in self.schema:
                        b = name.encode("utf-8")
                        buf += FIELD_DEF.to_bytes(1, "little")
                        buf += len(b).to_bytes(2, "little")
                        buf += b
                    self.headers_blob = bytes(buf)

                case "xlsx":
                    self.headers_blob = self.schema

                case _:
                    raise ValueError(f"Unsupported file type: {self.file_type}")

        except Exception as e:
            print(f"Exception in header: {e}")

    # =========================== PUBLISHER ========================
    def publish(self, values=None, encode=None):
        
        if not self._running or not self._enabled:
            return

        try:
            if values is None:
                raise ValueError("Values cannot be None")

            record = None
            if encode == None:
                encode = settings.ENCODER

            # =======================
            # 1. NORMALIZE INPUT    
            # =======================
            if isinstance(values, dict):
                if not self.schema:
                    raise RuntimeError(
                        "Schema not set. Call headers() before using dict input."
                    )
                processed_values = [values.get(k) for k in self.schema]

            elif isinstance(values, (list, tuple)):
                processed_values = list(values)

            elif isinstance(values, (bytes, bytearray)):
                processed_values = None  # raw binary path

            else:
                raise TypeError(
                    f"Unsupported input type: {type(values)}"
                )

            # =======================
            # 2. FORMAT HANDLING
            # =======================
            if self.file_type in ("bin", "tlv.bin"):

                # ---- RAW BINARY PATH ----
                if not encode:
                    if not isinstance(values, (bytes, bytearray)):
                        raise TypeError(
                            "encode=False requires raw bytes input or make it encode = True"
                        )
                    record = values

                # ---- STRUCTURED → BINARY ----
                else:
                    if processed_values is None:
                        raise TypeError(
                            "encode=True requires structured input (list/dict) or make it encode = False"
                        )

                    if self.file_type == "bin":
                        record = self._encode_record_bin(processed_values)
                    else:
                        record = self._encode_record_tlvbin(processed_values)

            elif self.file_type in ("csv", "xlsx"):
                if processed_values is None:
                    raise TypeError(
                        "CSV/XLSX do not accept raw binary input"
                    )
                record = processed_values

            else:
                raise ValueError(f"Unsupported file type: {self.file_type}")

            # =======================
            # 3. QUEUE
            # =======================
            try:
                self.q.put(record, timeout=0.001)
            except queue.Full:
                self.dropped_count += 1

        except Exception as e:
            self._enabled = False
            print(f"Exception in publish: {e}")

    # =========================== STOP ========================
    def stop(self):
        try:
            self._enabled = False
            self._running = False
        except Exception as e:
            print(f"Exception in stop: {e}")

    # =========================== ENCODERS ========================
    def _encode_record_bin(self, values):
        try:
            payload = bytearray()

            for value in values:

                if isinstance(value, bool):
                    payload += struct.pack("<?", value)

                elif isinstance(value, int):
                    payload += struct.pack("<q", value)   # int64

                elif isinstance(value, float):
                    payload += struct.pack("<d", value)   # float64

                elif isinstance(value, str):
                    payload += value.encode("utf-8") + b"\x00"  

                elif isinstance(value, bytes):
                    payload += value  # raw bytes (FIXED SIZE REQUIRED)

                else:
                    raise TypeError(f"Unsupported type: {type(value)}")

            return bytes(payload)
    
        except Exception as e:
            print(f"Exception in binary encoder: {e}")

    def _encode_record_tlvbin(self, values):
        try:
            if not self.schema:
                raise RuntimeError("Schema not set. Call headers() first.")

            if len(values) != len(self.schema):
                raise ValueError("Record does not match schema length")
            
            TYPE_BOOL   = 1
            TYPE_INT    = 2
            TYPE_FLOAT  = 3
            TYPE_STRING = 4
            TYPE_BYTES  = 5
            TYPE_NONE   = 6

            buf = bytearray()

            for value in values:

                # ---------- TYPE + VALUE ----------
                if value is None:
                    buf += TYPE_NONE.to_bytes(1, "little")
                    buf += (0).to_bytes(2, "little")

                elif isinstance(value, bool):
                    buf += TYPE_BOOL.to_bytes(1, "little")
                    buf += (1).to_bytes(2, "little")
                    buf += b"\x01" if value else b"\x00"

                elif isinstance(value, int):
                    data = struct.pack("<q", value)  # int64
                    buf += TYPE_INT.to_bytes(1, "little")
                    buf += len(data).to_bytes(2, "little")
                    buf += data

                elif isinstance(value, float):
                    data = struct.pack("<d", value)  # float64
                    buf += TYPE_FLOAT.to_bytes(1, "little")
                    buf += len(data).to_bytes(2, "little")
                    buf += data 

                elif isinstance(value, str):
                    data = value.encode("utf-8")
                    buf += TYPE_STRING.to_bytes(1, "little")
                    buf += len(data).to_bytes(2, "little")
                    buf += data

                elif isinstance(value, bytes):
                    buf += TYPE_BYTES.to_bytes(1, "little")
                    buf += len(value).to_bytes(2, "little")
                    buf += value

                else:
                    raise TypeError(f"Unsupported type: {type(value)}")

            record = bytearray()
            record += len(buf).to_bytes(2, "little")
            record += buf

            return bytes(record)

        except Exception as e:
            raise RuntimeError(f"Exception in TLV encoder: {e}")



        
    # =========================== BIN WORKER ========================
    def bin_worker(self):
        try:
            max_bytes = settings.MAX_FILE_SIZE_MB * 1024 * 1024
            current_size = 0
            # start = time.time()
            # sec_count = 0

            # open first file
            f = open(self.file_manager.current_file, "ab")

            # WRITE HEADER 
            f.write(self.headers_blob)
            f.flush()

            current_size = len(self.headers_blob)

            try:
                while self._running or not self.q.empty():
                    # end = time.time()
                    try:
                        record = self.q.get(timeout=0.1)
                    except queue.Empty:
                        continue

                    size = len(record)

                    if current_size + size >= max_bytes:
                        f.close()

                        self.file_manager.current_file = self.file_manager._new_log_file()
                        f = open(self.file_manager.current_file, "ab")
                        current_size = 0
                        self.file_no += 1

                        if self.headers_blob:
                            f.write(self.headers_blob)
                            f.flush()
                            current_size += len(self.headers_blob)

                        self.file_manager.compress_logs()


                    f.write(record)
                    current_size += size
                    # sec_count += 1

                    # if end - start >= 1.0:
                    #     print(f"[Worker] Exc: {sec_count}")
                    #     sec_count = 0
                    #     start = end

                    self.q.task_done()

            finally:
                f.close()
        except Exception as e:
            print(f"Exception in Bin Worker: {e}")

    # =========================== TLV BIN WORKER ========================
    def tlv_worker(self):
        try:
            max_bytes = settings.MAX_FILE_SIZE_MB * 1024 * 1024
            current_size = 0
            # start = time.time()
            # sec_count = 0

            f = open(self.file_manager.current_file, "ab")

            f.write(self.headers_blob)
            f.flush()

            current_size = len(self.headers_blob)

            try:
                while self._running or not self.q.empty():
                    # end = time.time()
                    try:
                        record = self.q.get(timeout=0.1)
                    except queue.Empty:
                        continue

                    size = len(record)

                    if current_size + size >= max_bytes:
                        f.close()

                        self.file_manager.current_file = self.file_manager._new_log_file()
                        f = open(self.file_manager.current_file, "ab")
                        current_size = 0
                        self.file_no += 1

                        if self.headers_blob:
                            f.write(self.headers_blob)
                            f.flush()
                            current_size += len(self.headers_blob)

                        self.file_manager.compress_logs()

            
                    f.write(record)
                    current_size += size
                    # sec_count += 1

                    # if end - start >= 1.0:
                    #     print(f"[Worker] Exc: {sec_count}")
                    #     sec_count = 0
                    #     start = end

                    self.q.task_done()

            finally:
                f.close()
        except Exception as e:
            print(f"Exception in TLV Bin Worker: {e}")


    # =========================== CSV WORKER ========================
    def csv_worker(self):
        try:
            max_bytes = settings.MAX_FILE_SIZE_MB * 1024 * 1024
            current_size = 0

            f = open(self.file_manager.current_file, "ab")

            if self.headers_blob:
                f.write(self.headers_blob)
                f.flush()
                current_size += len(self.headers_blob)

            try:
                while self._running or not self.q.empty():
                    try:
                        record = self.q.get(timeout=0.1)
                    except queue.Empty:
                        continue

                    line = ",".join(map(str, record)) + "\n"

                    encoded = line.encode("utf-8")
                    size = len(encoded)

                    if current_size + size >= max_bytes:
                        f.close()

                        self.file_manager.current_file = self.file_manager._new_log_file()
                        f = open(self.file_manager.current_file, "ab")
                        current_size = 0

                        if self.headers_blob:
                            f.write(self.headers_blob)
                            f.flush()
                            current_size += len(self.headers_blob)

                        self.file_manager.compress_logs()

                    f.write(encoded)
                    current_size += size
                    self.q.task_done()

            finally:
                f.close()
        except Exception as e:
            print(f"Exception in CSV worker: {e}")


    def xlsx_worker(self):
        try:
            # Excel hard limit ≈ 1,048,576
            MAX_ROWS = settings.XLSX_MAX_ROWS
            wb = Workbook(write_only=True)
            ws = wb.create_sheet(title="log")
            row_count = 0

            def prepare_new_sheet(workbook):
                sheet = workbook.create_sheet(title="log")
                count = 0
                if self.schema:
                    sheet.append(list(self.schema))
                    count = 1
                return sheet, count

            # Initial header setup
            if self.schema:
                ws.append(list(self.schema))
                row_count = 1

            try:
                while self._running or not self.q.empty():
                    try:
                        # Use a slightly longer timeout to reduce CPU spikes
                        record = self.q.get(timeout=0.5)
                    except queue.Empty:
                        continue

                    try:
                        ws.append(list(record))
                        row_count += 1
                    finally:
                        # Always mark task done even if append fails
                        self.q.task_done()

                    # Rotate XLSX file
                    if row_count >= MAX_ROWS:
                        wb.save(self.file_manager.current_file)
                        # self._compress_event.set()

                        # Setup new workbook
                        self.file_manager.current_file = self.file_manager._new_log_file()
                        wb = Workbook(write_only=True)
                        ws, row_count = prepare_new_sheet(wb)
                        self.file_no += 1

                        self.file_manager.compress_logs()


            except Exception as e:
                # Log your error here so the thread doesn't die silently
                print(f"Worker error: {e}")
            finally:
                # Only save if we actually wrote data beyond the header
                # or if the file doesn't exist yet.
                try:
                    wb.save(self.file_manager.current_file)
                    wb.close()
                except Exception:
                    pass

        except Exception as e:
            print(f"Exception in XLSX Worker: {e}")
